#!/usr/bin/env python3
"""
DCC Test Suite — Dark Echoes synthetic test files
Mirrors the JS logic from dubbing-clip-checker.html
"""
import sys, unicodedata, re
import openpyxl

BASE = "/Users/saidroblesnajera/Documents/dubbing-clip-checker/test_files"
PASS = "\033[32mPASS\033[0m"
FAIL = "\033[31mFAIL\033[0m"

results = []

def check(name, cond, detail=""):
    status = PASS if cond else FAIL
    msg = f"  [{status}] {name}"
    if detail:
        msg += f"\n         {detail}"
    print(msg)
    results.append((name, cond))

# ── Utils ────────────────────────────────────────────────────────────────────

def norm(s):
    """Mirrors JS norm(): NFD → strip combining → upper → strip apostrophes → spaces/hyphens → _"""
    s = unicodedata.normalize("NFD", s)
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    s = s.upper()
    s = re.sub(r"['\u2019`]", "", s)
    s = re.sub(r"[\s\-]+", "_", s)
    return s

def tc_to_frames(tc, fps=24):
    """Mirrors JS tcToFrames(): handles both : and ; separators"""
    if not tc:
        return None
    tc = tc.strip().replace(";", ":")
    parts = tc.split(":")
    if len(parts) < 4:
        return None
    h, m, s, f = (int(x) if x.isdigit() else 0 for x in parts)
    return h * 3600 * fps + m * 60 * fps + s * fps + f

def tc_overlap(a, b, c, d, t=2):
    if any(v is None for v in [a, b, c, d]):
        return False
    return a <= d + t and b >= c - t

def edit_dist(a, b):
    if len(a) > len(b):
        a, b = b, a
    r = list(range(len(a) + 1))
    for j in range(1, len(b) + 1):
        prev = r[0]
        r[0] = j
        for i in range(1, len(a) + 1):
            temp = r[i]
            r[i] = prev if a[i-1] == b[j-1] else 1 + min(prev, r[i], r[i-1])
            prev = temp
    return r[len(a)]

# ── Parse PT ─────────────────────────────────────────────────────────────────

def parse_pt(text, fps=24):
    clips = []
    lines = text.splitlines()
    in_tl = False
    track = ""
    in_clips = False
    expect_header = False

    for line in lines:
        if re.search(r"T\s*R\s*A\s*C\s*K\s+L\s*I\s*S\s*T\s*I\s*N\s*G", line, re.I):
            in_tl = True
            continue
        if not in_tl:
            continue
        if line.startswith("TRACK NAME:\t"):
            track = line.replace("TRACK NAME:\t", "").strip()
            expect_header = True
            in_clips = False
            continue
        if expect_header and re.search(r"CHANNEL\s.*EVENT\s.*CLIP NAME", line):
            expect_header = False
            in_clips = True
            continue
        if in_clips and line.strip() == "":
            in_clips = False
            continue
        if in_clips:
            parts = line.split("\t")
            if len(parts) >= 7:
                cn = parts[2].strip()
                st = parts[3].strip()
                et = parts[4].strip()
                if st and re.search(r"\d{2}[;:]\d{2}[;:]\d{2}[;:]\d{2}", st):
                    sf = tc_to_frames(st, fps)
                    ef = tc_to_frames(et, fps)
                    if sf is not None and ef is not None:
                        clips.append({
                            "track": track, "name": cn,
                            "tc_in": st, "tc_out": et,
                            "frames_in": sf, "frames_out": ef,
                            "duration": ef - sf
                        })

    # Filter stems > 2 min (2880 frames at 24fps = 120s)
    all_clips = clips[:]
    filtered = [c for c in clips if c["duration"] <= 2880]
    earliest = min((c["frames_in"] for c in all_clips), default=0) if all_clips else 0

    return filtered, all_clips, {"earliest_frame": earliest, "fps": fps}

# ── Parse PLDL ───────────────────────────────────────────────────────────────

def parse_pldl(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    # Find "Dialogue List" sheet or first
    sheet = None
    for name in wb.sheetnames:
        if re.search(r"dialogue\s*list", name, re.I):
            sheet = wb[name]
            break
    if not sheet:
        sheet = wb.active

    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []

    # Find header row with TC/TIMECODE
    header_idx = 0
    for i, row in enumerate(rows[:10]):
        if any(cell and re.search(r"timecode|tc\s*in|in.?timecode", str(cell), re.I) for cell in row):
            header_idx = i
            break

    headers = [str(h or "").strip().lower().replace(" ", "").replace("-", "").replace("_", "") for h in rows[header_idx]]

    def find(pats):
        for i, h in enumerate(headers):
            if any(p in h for p in pats):
                return i
        return -1

    tc_in_i = find(["timecodein", "intimecode", "tcin", "timein"])
    tc_out_i = find(["timecodeout", "outtimecode", "tcout", "timeout"])
    src_i    = find(["charactersource", "source", "character", "personaje"])
    transc_i = find(["transcription", "transcripcion", "contenttranscription"])
    dialog_i = find(["dialogue", "dialog", "parlamento"])
    tag_i    = find(["tags", "tag", "type"])

    entries = []
    idx = 0
    for row in rows[header_idx + 1:]:
        tc_in = str(row[tc_in_i] or "").strip() if tc_in_i >= 0 else ""
        if not tc_in or not re.search(r"\d{2}[;:]\d{2}[;:]\d{2}[;:]\d{2}", tc_in):
            continue
        idx += 1
        tc_out = str(row[tc_out_i] or "").strip() if tc_out_i >= 0 else ""
        source = str(row[src_i] or "").strip() if src_i >= 0 else ""
        transc = str(row[transc_i] or "").strip() if transc_i >= 0 else ""
        dialog = str(row[dialog_i] or "").strip() if dialog_i >= 0 else ""
        tags   = str(row[tag_i] or "").strip() if tag_i >= 0 else ""
        entries.append({
            "idx": idx, "source": source, "transcription": transc,
            "dialog": dialog, "tags": tags,
            "tc_in": tc_in, "tc_out": tc_out,
            "frames_in": tc_to_frames(tc_in), "frames_out": tc_to_frames(tc_out)
        })
    return entries

# ── buildSTMap ────────────────────────────────────────────────────────────────

def build_st_map(entries, clips):
    tracks = list(dict.fromkeys(c["track"] for c in clips))
    sources = list(dict.fromkeys(e["source"] for e in entries if e["source"]))
    m = {}
    for src in sources:
        sn = norm(src)
        sf = norm(src.split()[0]) if src.split() else sn
        # Exact match
        t = next((t for t in tracks if
                  norm(t).startswith(sn + "_") or norm(t) == sn or
                  norm(t).startswith(sf + "_") or norm(t) == sf), None)
        if t:
            m[src] = t
            continue
        # Fuzzy (edit distance ≤ 2)
        best = None
        best_d = 3
        for tr in tracks:
            tn = norm(tr)
            prefix = "_".join(tn.split("_")[:len(sn.split("_"))])
            d = edit_dist(sn, prefix)
            if d < best_d:
                best_d = d
                best = tr
        if best:
            m[src] = best
    return m

# ── matchClips ────────────────────────────────────────────────────────────────

def belongs(clip_name, src):
    if not src:
        return False
    sn = norm(src)
    sf = norm(src.split()[0]) if src.split() else sn
    nn = norm(clip_name)
    if sn in nn or nn.startswith(sf + "_"):
        return True
    prefix = "_".join(nn.split("_")[:len(sn.split("_"))])
    return edit_dist(sn, prefix) <= 2

def match_clips(entries, clips, tol, st_map, offset=0):
    tcm = {}
    for c in clips:
        tcm.setdefault(c["track"], []).append(c)

    results_list = []
    for e in entries:
        f_in = e["frames_in"] + offset
        f_out = e["frames_out"] + offset
        tn = st_map.get(e["source"])
        if tn:
            cands = tcm.get(tn, [])
            m = next((c for c in cands if tc_overlap(f_in, f_out, c["frames_in"], c["frames_out"], tol)), None)
            if m:
                results_list.append({**e, "matched": True, "matched_clip": m})
                continue
        # Global search
        ext = next((c for c in clips if
                    (not tn or c["track"] != tn) and
                    belongs(c["name"], e["source"]) and
                    tc_overlap(f_in, f_out, c["frames_in"], c["frames_out"], tol)), None)
        if ext:
            results_list.append({**e, "matched": True, "matched_clip": ext})
        else:
            results_list.append({**e, "matched": False, "matched_clip": None})
    return results_list

# ── KNP word-boundary check ───────────────────────────────────────────────────

def is_letter(ch):
    return bool(re.match(r"[a-zA-Z\u00C0-\u024F\u0400-\u04FF]", ch))

def knp_matches(text, term):
    """Returns True if term appears in text with word boundaries on both sides."""
    tl = term.lower()
    txt = text.lower()
    pos = txt.find(tl)
    while pos != -1:
        before = pos == 0 or not is_letter(txt[pos - 1])
        after = pos + len(tl) >= len(txt) or not is_letter(txt[pos + len(tl)])
        if before and after:
            return True
        pos = txt.find(tl, pos + 1)
    return False

# ════════════════════════════════════════════════════════════════════════════
# TESTS
# ════════════════════════════════════════════════════════════════════════════

print("\n\033[1m── DCC Test Suite — Dark Echoes ──\033[0m\n")

# ── 1. Parse PLDL ────────────────────────────────────────────────────────────
print("\033[1m1. PLDL Parsing\033[0m")
pldl_entries = parse_pldl(f"{BASE}/test_PLDL.xlsx")

check("36 total entries parsed (README says 35 — file has 36 rows)",
      len(pldl_entries) == 36,
      f"got {len(pldl_entries)}")

# JS SKIP list uses substring includes() — "AD" in "BRADLEY" → NICK/VANESSA BRADLEY filtered
# This is actual app behavior, though unintended for BRADLEY names
SKIP_SOURCES = ["WALLA", "MAIN TITLE", "GRAPHICS INSERTS", "GRAPHICS", "AD",
                "FOREST BLAKK", "JOY WILLIAMS", "PRINCIPAL PHOTOGRAPHY"]
SKIP_TAGS = ["MAIN TITLE", "GRAPHICS INSERTS", "INAUDIBLE", "SONG"]

def should_exclude(e):
    """Mirrors fixed JS filter: s===sk || s.startsWith(sk+" ")"""
    s = e["source"].upper()
    if any(s == sk or s.startswith(sk + " ") for sk in SKIP_SOURCES):
        return True
    tags = [t.strip().upper() for t in e["tags"].split(",")]
    return any(tag in SKIP_TAGS for tag in tags)

filtered_entries = [e for e in pldl_entries if not should_exclude(e)]
bradley_filtered = [e for e in pldl_entries if "BRADLEY" in e["source"].upper() and e not in filtered_entries]
check("31 entries after non-dialog filter (NICK/VANESSA BRADLEY now correctly kept)",
      len(filtered_entries) == 31,
      f"got {len(filtered_entries)}; BRADLEY entries in filtered: {len(bradley_filtered)}")

# ── 2. Parse Pro Tools ───────────────────────────────────────────────────────
print("\n\033[1m2. Pro Tools Session Parsing\033[0m")
with open(f"{BASE}/test_ProTools_SessionInfo.txt", encoding="latin-1") as f:
    pt_text = f.read()

pt_clips, all_pt_clips, pt_meta = parse_pt(pt_text, fps=24)

stems = [c for c in all_pt_clips if c["duration"] > 2880]
check("Stems > 2 min detected",
      any("DIA_STEM" in c["name"] or "PM_OV" in c["name"] or "LAS_MIX" in c["name"] for c in stems),
      f"stems found: {[c['name'] for c in stems]}")

check("Stems filtered out of working clips",
      not any(c["duration"] > 2880 for c in pt_clips),
      f"max duration in filtered: {max((c['duration'] for c in pt_clips), default=0)} frames")

check("Valid clips remain after filtering",
      len(pt_clips) > 0,
      f"{len(pt_clips)} clips kept")

# ── 3. Clip Matching ─────────────────────────────────────────────────────────
print("\n\033[1m3. Clip Matching\033[0m")
st_map = build_st_map(filtered_entries, pt_clips)
match_results = match_clips(filtered_entries, pt_clips, tol=6, st_map=st_map, offset=0)

# NARRATOR entries: identify them for exclusion check
narrator_entries = [r for r in match_results if r["source"].upper() == "NARRATOR"]
non_narrator = [r for r in match_results if r["source"].upper() != "NARRATOR"]

matched = [r for r in non_narrator if r["matched"]]
missing = [r for r in non_narrator if not r["matched"]]

check("27 grabados (non-NARRATOR matched)",
      len(matched) == 27,
      f"got {len(matched)}; missing: {[(r['source'], r['tc_in'], r['dialog'][:40]) for r in missing]}")

check("2 faltantes (Elena 02:40 + Marco 02:44)",
      len(missing) == 2 and
      any("02:40" in r["tc_in"] and "ELENA" in r["source"].upper() for r in missing) and
      any("02:44" in r["tc_in"] and "MARCO" in r["source"].upper() for r in missing),
      f"missing: {[(r['source'], r['tc_in']) for r in missing]}")

check("2 NARRATOR entries (excluible)",
      len(narrator_entries) == 2,
      f"got {len(narrator_entries)}")

# ── 4. Edge Cases ─────────────────────────────────────────────────────────────
print("\n\033[1m4. Edge Cases\033[0m")

# 4a. Apostrophe normalization
_zack_norm = norm("ZACK'S MOTHER")
check("norm: ZACK'S MOTHER → ZACKS_MOTHER",
      _zack_norm == "ZACKS_MOTHER",
      f"got: {_zack_norm}")

zacks_entry = next((e for e in filtered_entries if "ZACK" in e["source"].upper()), None)
zacks_result = next((r for r in match_results if zacks_entry and r["idx"] == zacks_entry["idx"]), None)
check("ZACK'S MOTHER matched in PT (apostrophe normalization)",
      zacks_result is not None and zacks_result["matched"],
      f"result: {zacks_result}")

# 4b. Diacritic normalization
check("norm: MÜLLER → MULLER",
      norm("MÜLLER") == "MULLER",
      f"got: {norm('MÜLLER')}")

muller_entry = next((e for e in filtered_entries if "LLER" in e["source"].upper() or "ÜLLER" in e["source"].upper()), None)
muller_result = next((r for r in match_results if muller_entry and r["idx"] == muller_entry["idx"]), None)
check("MÜLLER matched in PT (diacritic normalization)",
      muller_result is not None and muller_result["matched"],
      f"source: {muller_entry['source'] if muller_entry else 'NOT FOUND'}, result: {muller_result}")

# 4c. INSPECTOR GRAF in NICK_BRADLEY track
graf_entries = [e for e in filtered_entries if "INSPECTOR" in e["source"].upper() and "GRAF" in e["source"].upper()]
graf_results  = [r for r in match_results if r["source"] in [e["source"] for e in graf_entries]]
graf_cross = [r for r in graf_results if r["matched"] and r["matched_clip"] and "NICK_BRADLEY" in r["matched_clip"]["track"]]
check("INSPECTOR GRAF clips found in NICK_BRADLEY track",
      len(graf_cross) > 0,
      f"graf results: {[(r['tc_in'], r['matched_clip']['track'] if r['matched_clip'] else 'NONE') for r in graf_results]}")

# 4d. ELENA VOSS phone clip in FUTZ track
elena_futz = [r for r in match_results
              if r["matched"] and r["matched_clip"] and
              "FUTZ" in r["matched_clip"]["track"].upper() and
              "ELENA" in r["source"].upper()]
check("ELENA VOSS phone clip matched in FUTZ PHONE track",
      len(elena_futz) > 0,
      f"futz matches: {[(r['tc_in'], r['matched_clip']['name']) for r in elena_futz]}")

# 4e. Tag "MAIN TITLE IN DIALOGUE,CONTEXT" NOT excluded
tag_entries = [e for e in pldl_entries if "MAIN TITLE IN DIALOGUE" in e["tags"].upper()]
check("Tag 'MAIN TITLE IN DIALOGUE,CONTEXT' NOT excluded (exact match only)",
      len(tag_entries) > 0 and all(e in filtered_entries for e in tag_entries),
      f"found {len(tag_entries)} such entries, in filtered: {sum(1 for e in tag_entries if e in filtered_entries)}")

# 4f. Drop frame TCs (00:03:10;05) parsed correctly
df_entries = [e for e in pldl_entries if ";" in e["tc_in"]]
check("Drop frame TCs (;) parsed correctly",
      len(df_entries) >= 2 and all(e["frames_in"] is not None for e in df_entries),
      f"drop frame entries: {[(e['tc_in'], e['frames_in']) for e in df_entries]}")

# ── 5. KNP Word Boundary ─────────────────────────────────────────────────────
print("\n\033[1m5. KNP Word Boundary\033[0m")

check("'Vo' does NOT match inside 'Voss'",
      not knp_matches("Elena Voss", "Vo"),
      "'Vo' incorrectly matched inside 'Voss'")

check("'Voss' DOES match as standalone word",
      knp_matches("Elena Voss was here", "Voss"),
      "'Voss' did not match standalone")

check("'the coin' matches in context",
      knp_matches("She found the coin on the table", "the coin"),
      "'the coin' did not match")

check("'coin' does NOT match inside 'coincidence'",
      not knp_matches("What a coincidence", "coin"),
      "'coin' incorrectly matched inside 'coincidence'")

# ── Summary ───────────────────────────────────────────────────────────────────
total = len(results)
passed = sum(1 for _, ok in results if ok)
failed = total - passed
print(f"\n\033[1m── Results: {passed}/{total} passed", end="")
if failed:
    print(f" (\033[31m{failed} failed\033[0m\033[1m)", end="")
print(" ──\033[0m")
if failed:
    print("\nFailed checks:")
    for name, ok in results:
        if not ok:
            print(f"  • {name}")

sys.exit(0 if failed == 0 else 1)
